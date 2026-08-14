"""
market_momentum.py - 全市场动量观察模块（外部观察信号，非生产决策）

来源：TheMarketMemo「全市场动量观察表」（tradecat 公开工作簿，导出 xlsx）。
定位：本模块只做「摄入 -> 复刻计算 -> 观察输出 -> 累积时间序列」，
      绝不进入生产决策链、不修改 risk_guard / shadow / run_daily。
      它是日报里的「外部参考」一节，以及未来回测（P2）的数据底座。

====================================================================
真实方法论（从导出 xlsx 的 GOOGLEFINANCE 公式还原，可信度：高）
====================================================================
对全宇宙（资产/市场结构/行业 三 tab 合并约 219 只）每只标的、每个窗口 W：
  1) 超额收益 excess(W) = 标的 W 日收益 - SPY 的 W 日收益        （基准 = SPY）
  2) 相对强度  RS(W)     = (1 - excess(W) 在宇宙内的百分位排名) * 100   （0~100）
  3) 综合排名  composite  = 加权求和 RS(W)
  4) 锚点收益  anchor_ret = 自固定锚点日（导出表为 2025-12-31）至今累计收益

=> 标准「相对动量百分位 + 加权合成」框架：只排名、不预测。
   与系统北极星「Global Multi-Asset CIO / 动态管理风险暴露」一致。

====================================================================
模型预设（Model A/B 当前可算；C/D/E 待数据）
====================================================================
  Model A (baseline) : 0.20*RS20 + 0.40*RS60 + 0.40*RS120   —— 与源表公式一致（drift=0 已验证）
  Model B (加 5D)    : 0.10*RS5  + 0.20*RS20 + 0.35*RS60 + 0.35*RS120
                       RS5 由本模块用「全宇宙 5 日超额收益百分位」独立计算（表内已有 5 日超额列）
  Model C (>200DMA)  : 需价格序列（表内无）=> 占位，未实现
  Model D (动量/波动): 需波动率序列（表内无）=> 占位，未实现
  Model E (Regime 调制): 需 Regime 信号（接 risk_governance）=> 占位，未实现

====================================================================
用法
====================================================================
  python market_momentum.py [动量表路径]   # 接受 .xlsx 或 .csv（Google Sheet /export?format=csv 直链）
  默认路径：D:/Downloads/全市场动量观察表.xlsx
  也接受：backend/imports/market_momentum_latest.xlsx 或 .csv
  每日自动化由 fetch_market_sheets.py 下载 CSV 后调用本模块（同日快照自动去重）
输出：
  backend/output/momentum_latest.json              （最新快照，含 Model A/B/Tilt）
  backend/output/momentum_timeseries.jsonl        （时间序列，每次运行追加一行，P1 数据底座；
                                                    每行含 compA/compB/tilt 及 ft_ret_1w/2w/4w 占位列）
  backend/output/market_momentum_<数据日期>.html  （可读报告，含 Tilt 观察卡）
  backend/output/market_momentum_<数据日期>.md    （日报引用文本）

====================================================================
研究指标：Momentum Tilt = Model B − Model A
====================================================================
  Tilt 捕捉「中长期趋势(A) 之上叠加的短期动量倾斜(B)」，回答：
    “加入短期动量后，能否更早识别趋势延续(A高+Tilt高)与趋势衰减(A高+Tilt低)？”
  五档观察阈值（研究用，非交易信号）：
    >+10 显著增强 ｜ +5~+10 改善 ｜ −5~+5 稳定 ｜ −10~−5 走弱 ｜ <−10 明显恶化
  预测价值需 P2（≥20 周度快照）做 Incremental Information Test 验证。
"""

import os
import sys
import json
import bisect
import datetime as dt
from collections import defaultdict

import openpyxl

# ----------------------------------------------------------------------------
# 配置
# ----------------------------------------------------------------------------
HERE = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.abspath(os.path.join(HERE, "..", "output"))
IMPORTS_DIR = os.path.abspath(os.path.join(HERE, "..", "imports"))
TIMESERIES_PATH = os.path.join(OUTPUT_DIR, "momentum_timeseries.jsonl")
DEFAULT_XLSX = r"D:/Downloads/全市场动量观察表.xlsx"
DATA_TABS = ["资产", "市场结构", "行业"]   # 三个业务数据表

# 模型预设：weights 为各窗口 RS 权重；need_rs5 决定是否计算 RS5
MODELS = {
    "A": {"label": "Baseline（源表公式）", "weights": {20: 0.20, 60: 0.40, 120: 0.40}, "need_rs5": False},
    "B": {"label": "加 5D 短期动量", "weights": {5: 0.10, 20: 0.20, 60: 0.35, 120: 0.35}, "need_rs5": True},
}
PRIMARY_MODEL = "A"   # 轮动信号主模型 = 已验证基准

# 分类阈值（启发式，非源表逻辑，仅用于观察分组）
STRONG, NEUTRAL = 70.0, 40.0


# ----------------------------------------------------------------------------
# 1) 摄入：读取 xlsx
# ----------------------------------------------------------------------------
def _first_nonempty_row(ws):
    for r in ws.iter_rows(values_only=True):
        if any(c is not None for c in r):
            return r
    return None


def _find_idx(header, name):
    for i, h in enumerate(header):
        if h == name:
            return i
    return None


def parse_workbook(path):
    """读取动量表，返回 {tab: [record, ...]}。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    for tab in DATA_TABS:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        rows = list(ws.iter_rows(values_only=True))
        hdr = None
        start = 0
        for i, r in enumerate(rows):
            if any(c is not None for c in r):
                hdr = list(r)
                start = i + 1
                break
        if hdr is None:
            continue

        out[tab] = _build_tab_recs(hdr, rows[start:])
    wb.close()
    return out


def _build_tab_recs(hdr, data_rows):
    """从表头 + 数据行构建记录列表（xlsx 与 csv 共用同一套列定位逻辑）。"""
    i_cat1 = _find_idx(hdr, "一级分类")
    i_cat2 = _find_idx(hdr, "二级分类")
    i_tkr = _find_idx(hdr, "标的代码")
    i_cn = _find_idx(hdr, "中文名称")
    i_en = _find_idx(hdr, "英文名称")
    i_px = _find_idx(hdr, "价格")
    i_dchg = _find_idx(hdr, "单日涨跌")
    i_rs20 = _find_idx(hdr, "20日相对强度")
    i_rs60 = _find_idx(hdr, "60日相对强度")
    i_rs120 = _find_idx(hdr, "120日相对强度")
    i_comp = _find_idx(hdr, "综合排名")
    i_ttime = _find_idx(hdr, "交易时间")
    if i_comp is None:
        return []
    i_ex5, i_ex20, i_ex60, i_ex120 = i_comp + 1, i_comp + 2, i_comp + 3, i_comp + 4
    i_anchor = i_comp + 5

    def num(x):
        try:
            return float(x) if x is not None else None
        except (TypeError, ValueError):
            return None

    recs = []
    for r in data_rows:
        if len(r) <= i_tkr or r[i_tkr] in (None, ""):
            continue
        rec = {
            "cat1": r[i_cat1] if i_cat1 is not None else None,
            "cat2": r[i_cat2] if i_cat2 is not None else None,
            "ticker": str(r[i_tkr]).strip(),
            "cn_name": r[i_cn] if i_cn is not None else None,
            "en_name": r[i_en] if i_en is not None else None,
            "price": num(r[i_px]) if i_px is not None else None,
            "daily_chg": num(r[i_dchg]) if i_dchg is not None else None,
            "rs20": num(r[i_rs20]) if i_rs20 is not None else None,
            "rs60": num(r[i_rs60]) if i_rs60 is not None else None,
            "rs120": num(r[i_rs120]) if i_rs120 is not None else None,
            "composite_src": num(r[i_comp]) if i_comp is not None else None,
            "ex5": num(r[i_ex5]) if i_ex5 < len(r) else None,
            "ex20": num(r[i_ex20]) if i_ex20 < len(r) else None,
            "ex60": num(r[i_ex60]) if i_ex60 < len(r) else None,
            "ex120": num(r[i_ex120]) if i_ex120 < len(r) else None,
            "anchor_ret": num(r[i_anchor]) if i_anchor < len(r) else None,
            "trade_time": str(r[i_ttime]) if (i_ttime is not None and i_ttime < len(r) and r[i_ttime] is not None) else None,
        }
        recs.append(rec)
    return recs


def parse_csv(path):
    """读取导出的 CSV（Google Sheet /export?format=csv 直链），返回 {'CSV': [record, ...]}。

    CSV 列名与 xlsx 一致，复用 _build_tab_recs 的列定位逻辑。
    """
    import csv
    with open(path, newline='', encoding='utf-8-sig') as f:
        rows = [list(r) for r in csv.reader(f)]
    hdr = None
    start = 0
    for i, r in enumerate(rows):
        if any(c is not None and str(c).strip() != '' for c in r):
            hdr = [(c.strip() if isinstance(c, str) else c) for c in r]
            start = i + 1
            break
    if hdr is None:
        return {}
    return {"CSV": _build_tab_recs(hdr, rows[start:])}


def parse_any(path):
    """按扩展名分发：.csv 走 parse_csv，其余走 parse_workbook(xlsx)。"""
    if str(path).lower().endswith(".csv"):
        return parse_csv(path)
    return parse_workbook(path)


# ----------------------------------------------------------------------------
# 2) RS5 计算 + 多模型 composite
# ----------------------------------------------------------------------------
def compute_rs5(all_recs):
    """用全宇宙 5 日超额收益的百分位排名计算 RS5 = (1 - rank/N)*100。

    源表 RS 是在「全宇宙」内做横截面排名；RS5 同法处理，保证口径一致。
    返回 {ticker: rs5 or None}。
    """
    ex5_list = [r["ex5"] for r in all_recs if r.get("ex5") is not None]
    if not ex5_list:
        return {r["ticker"]: None for r in all_recs}
    srt = sorted(ex5_list)
    N = len(srt)
    out = {}
    for r in all_recs:
        e = r.get("ex5")
        if e is None:
            out[r["ticker"]] = None
            continue
        le = bisect.bisect_right(srt, e)   # 小于等于 e 的个数 = 排名
        rs5 = (1.0 - le / N) * 100.0
        out[r["ticker"]] = round(rs5, 6)
    return out


def model_composite(rec, model_key, rs5_map):
    """按预设模型算 composite；缺任一所需 RS 则返回 None。"""
    cfg = MODELS[model_key]
    w = cfg["weights"]
    if cfg["need_rs5"]:
        rs5 = rs5_map.get(rec["ticker"])
        if rs5 is None:
            return None
    total = 0.0
    for win, weight in w.items():
        if win == 5:
            val = rs5_map.get(rec["ticker"])
        else:
            val = rec.get("rs" + str(win))
        if val is None:
            return None
        total += weight * val
    return round(total, 6)


# ----------------------------------------------------------------------------
# 3) 观察输出：分类 + 轮动信号 + 模型对比
# ----------------------------------------------------------------------------
def _classify(composite):
    if composite is None:
        return "未知"
    if composite >= STRONG:
        return "强动量"
    if composite >= NEUTRAL:
        return "中性"
    return "弱势"


def _tilt_classify(tilt):
    """Momentum Tilt 研究观察分类（Tilt = Model B - Model A）。

    注意：这是研究观察阈值，不是交易信号；预测价值需 P2（≥20 周度快照）统计验证。
    """
    if tilt is None:
        return "未知"
    if tilt > 10:
        return "短期显著增强"
    if tilt > 5:
        return "短期改善"
    if tilt >= -5:
        return "基本稳定"
    if tilt >= -10:
        return "短期走弱"
    return "短期明显恶化"


# Tilt 研究观察阈值（仅观察，非信号）
TILT_THRESHOLDS = {
    "> +10": "短期动量显著增强（趋势修复/加速）",
    "+5~+10": "短期改善",
    "-5~+5": "基本稳定",
    "-10~-5": "短期走弱",
    "< -10": "短期明显恶化（趋势衰减）",
}


def build_snapshot(parsed):
    all_recs = []
    for tab, recs in parsed.items():
        for r in recs:
            r = dict(r)
            r["_tab"] = tab
            all_recs.append(r)

    rs5_map = compute_rs5(all_recs)
    for r in all_recs:
        r["rs5"] = rs5_map.get(r["ticker"])
        r["compA"] = model_composite(r, "A", rs5_map)
        r["compB"] = model_composite(r, "B", rs5_map)
        r["tilt"] = round(r["compB"] - r["compA"], 6) if (r["compA"] is not None and r["compB"] is not None) else None
        r["tilt_cls"] = _tilt_classify(r["tilt"])
        # 源表 composite 复刻校验（仅 Model A 对应源表）
        rec, src, drift = _recompute_src(r)
        r["composite_recomputed"] = rec
        r["composite_drift"] = drift
        r["composite"] = r["compA"] if r["compA"] is not None else r.get("composite_src")
        r["tier"] = _classify(r["composite"])

    # 轮动信号（基于主模型 A）
    grp = defaultdict(list)
    for r in all_recs:
        if r.get("composite") is not None:
            grp[r["cat2"]].append(r["composite"])
    grp_avg = {k: round(sum(v) / len(v), 2) for k, v in grp.items()}
    grp_ranked = sorted(grp_avg.items(), key=lambda x: x[1], reverse=True)

    grp1 = defaultdict(list)
    for r in all_recs:
        if r.get("composite") is not None:
            grp1[r["cat1"]].append(r["composite"])
    grp1_avg = {k: round(sum(v) / len(v), 2) for k, v in grp1.items()}
    grp1_ranked = sorted(grp1_avg.items(), key=lambda x: x[1], reverse=True)

    watch, avoid = [], []
    for r in all_recs:
        c = r.get("composite")
        e120 = r.get("ex120")
        anc = r.get("anchor_ret")
        if c is None:
            continue
        if c >= 60 and (e120 is None or e120 > 0) and (anc is None or anc > 0):
            watch.append(r)
        elif c < NEUTRAL or (e120 is not None and e120 < 0 and (r.get("ex60") or 0) < 0):
            avoid.append(r)
    watch.sort(key=lambda r: r["composite"], reverse=True)
    avoid.sort(key=lambda r: (r["composite"] if r["composite"] is not None else 0))

    ranked = [r for r in all_recs if r.get("composite") is not None]
    ranked.sort(key=lambda r: r["composite"], reverse=True)
    top15 = ranked[:15]

    # 数据日期
    data_date = dt.date.today().isoformat()
    dates = []
    for r in all_recs:
        tt = r.get("trade_time")
        if tt:
            try:
                dates.append(dt.datetime.strptime(tt[:10], "%Y-%m-%d").date())
            except Exception:
                pass
    if dates:
        data_date = max(dates).isoformat()

    # 复刻漂移（A 应 ~0）
    drifts = [r["composite_drift"] for r in all_recs if r.get("composite_drift") is not None]
    drift_max = max(drifts) if drifts else None
    drift_min = min(drifts) if drifts else None

    # Model A vs B 敏感性：排名变动大的标的
    compB_map = {r["ticker"]: r["compB"] for r in all_recs if r.get("compB") is not None}
    rankA = {r["ticker"]: i + 1 for i, r in enumerate(ranked)}
    rankB_list = [r for r in all_recs if r.get("compB") is not None]
    rankB_list.sort(key=lambda r: r["compB"], reverse=True)
    rankB = {r["ticker"]: i + 1 for i, r in enumerate(rankB_list)}
    sensitivity = []
    for tkr in rankA:
        if tkr in rankB:
            delta = rankB[tkr] - rankA[tkr]   # 正=在 B 中排名下降（更弱）
            if abs(delta) >= 15:
                sensitivity.append({"ticker": tkr, "rankA": rankA[tkr], "rankB": rankB[tkr], "delta": delta})
    sensitivity.sort(key=lambda x: abs(x["delta"]), reverse=True)

    # Momentum Tilt 观察（Tilt = Model B - Model A，研究阈值非信号）
    tilt_records = [r for r in all_recs if r.get("tilt") is not None]
    tilt_up = sorted([r for r in tilt_records if r["tilt"] > 5], key=lambda r: r["tilt"], reverse=True)
    tilt_dn = sorted([r for r in tilt_records if r["tilt"] < -5], key=lambda r: r["tilt"])
    tilt_view = {
        "strongest_up": [
            {"ticker": r["ticker"], "cn": r["cn_name"], "cat2": r["cat2"],
             "compA": r["compA"], "compB": r["compB"], "tilt": r["tilt"], "cls": r["tilt_cls"]}
            for r in tilt_up[:8]
        ],
        "strongest_down": [
            {"ticker": r["ticker"], "cn": r["cn_name"], "cat2": r["cat2"],
             "compA": r["compA"], "compB": r["compB"], "tilt": r["tilt"], "cls": r["tilt_cls"]}
            for r in tilt_dn[:8]
        ],
        "thresholds": TILT_THRESHOLDS,
        "warning": "以上为研究观察阈值，非交易信号；预测价值需 P2（≥20 周度快照）统计验证",
    }

    snapshot = {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "data_date": data_date,
        "source": "全市场动量观察表.xlsx（TheMarketMemo / tradecat 公开工作簿）",
        "methodology": {
            "benchmark": "SPY",
            "rs_definition": "(1 - 超额收益在宇宙内百分位排名) * 100",
            "models": {k: {"label": v["label"], "formula": _formula_str(k)} for k, v in MODELS.items()},
            "primary_model": PRIMARY_MODEL,
            "note": "相对动量百分位，只排名不预测；Model C/D/E 待价格序列/波动率/Regime 信号接入",
        },
        "universe_size": len(all_recs),
        "tier_distribution": _tier_counts(all_recs),
        "rotation_signal": {
            "leading_sectors": grp_ranked[:3],
            "lagging_sectors": grp_ranked[-3:][::-1],
            "leading_asset_classes": grp1_ranked[:3],
            "lagging_asset_classes": grp1_ranked[-3:][::-1],
        },
        "watch_pool": [
            {"ticker": r["ticker"], "cn": r["cn_name"], "cat2": r["cat2"],
             "composite": r["composite"], "ex120": r["ex120"], "anchor_ret": r["anchor_ret"]}
            for r in watch[:20]
        ],
        "avoid_pool": [
            {"ticker": r["ticker"], "cn": r["cn_name"], "cat2": r["cat2"],
             "composite": r["composite"], "ex120": r["ex120"]}
            for r in avoid[:20]
        ],
        "top15": [
            {"ticker": r["ticker"], "cn": r["cn_name"], "cat2": r["cat2"],
             "composite": r["composite"], "compB": r["compB"],
             "rs20": r["rs20"], "rs60": r["rs60"], "rs120": r["rs120"], "rs5": r["rs5"],
             "ex120": r["ex120"], "anchor_ret": r["anchor_ret"]}
            for r in top15
        ],
        "model_sensitivity": sensitivity[:20],
        "tilt_view": tilt_view,
        "replication_drift": {
            "max": drift_max, "min": drift_min,
            "meaning": "Model A 重算 composite 与源表值的最大/最小偏差；应~0，证伪/验证摄入一致性",
        },
        "records": all_recs,
    }
    return snapshot


def _formula_str(model_key):
    w = MODELS[model_key]["weights"]
    parts = []
    for win in sorted(w.keys()):
        parts.append(str(w[win]) + "*RS" + str(win))
    return " + ".join(parts)


def _recompute_src(rec):
    rs20, rs60, rs120 = rec.get("rs20"), rec.get("rs60"), rec.get("rs120")
    if None in (rs20, rs60, rs120):
        return None, rec.get("composite_src"), None
    recomputed = 0.20 * rs20 + 0.40 * rs60 + 0.40 * rs120
    src = rec.get("composite_src")
    drift = None if src is None else round(recomputed - src, 6)
    return round(recomputed, 6), src, drift


def _tier_counts(recs):
    c = defaultdict(int)
    for r in recs:
        c[r.get("tier")] += 1
    return dict(c)


# ----------------------------------------------------------------------------
# 4) 时间序列累积（P1 数据底座）
# ----------------------------------------------------------------------------
def append_timeseries(snapshot):
    """每次运行追加一行到 momentum_timeseries.jsonl（按 data_date 去重，避免每日自动化产生重复快照）。

    每行 = {date, models:{A:n,B:n}, rows:[{ticker,cat2,compA,compB,tilt,tilt_cls,rs5,rs20,rs60,rs120,
                                          ex5,ex20,ex60,ex120,anchor_ret, ft_ret_1w, ft_ret_2w, ft_ret_4w}]}
    随运行次数累积，构成未来回测（P2）的时间序列底座。
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    # 同 data_date 去重：仅当该日期尚无快照时才追加（每日自动化安全，不撑爆底座）
    if os.path.exists(TIMESERIES_PATH):
        try:
            with open(TIMESERIES_PATH, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line and json.loads(line).get("date") == snapshot["data_date"]:
                        print("      [去重] 该 data_date 已存在，跳过追加：" + snapshot["data_date"])
                        return TIMESERIES_PATH
        except Exception:
            pass
    row = {
        "date": snapshot["data_date"],
        "generated_at": snapshot["generated_at"],
        "models": {k: _formula_str(k) for k in MODELS},
        "rows": [],
    }
    for r in snapshot["records"]:
        row["rows"].append({
            "ticker": r["ticker"],
            "cat2": r.get("cat2"),
            "compA": r.get("compA"),
            "compB": r.get("compB"),
            "tilt": r.get("tilt"),
            "tilt_cls": r.get("tilt_cls"),
            "rs5": r.get("rs5"),
            "rs20": r.get("rs20"),
            "rs60": r.get("rs60"),
            "rs120": r.get("rs120"),
            "ex5": r.get("ex5"),
            "ex20": r.get("ex20"),
            "ex60": r.get("ex60"),
            "ex120": r.get("ex120"),
            "anchor_ret": r.get("anchor_ret"),
            # P2 研究占位列：未来收益（待接入价格序列后回填，当前为 None）
            "ft_ret_1w": None,
            "ft_ret_2w": None,
            "ft_ret_4w": None,
        })
    with open(TIMESERIES_PATH, "a", encoding="utf-8") as f:
        f.write(json.dumps(row, ensure_ascii=False, default=str) + "\n")
    return TIMESERIES_PATH


def load_timeseries():
    """读取全部历史快照，返回按日期排序的 list。空文件返回 []。"""
    if not os.path.exists(TIMESERIES_PATH):
        return []
    out = []
    with open(TIMESERIES_PATH, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                out.append(json.loads(line))
    return out


def timeseries_summary():
    """给报告用的轻量时间序列摘要（最新 vs 首次，含样本数）。"""
    hist = load_timeseries()
    if not hist:
        return None
    return {
        "snapshots": len(hist),
        "first_date": hist[0]["date"],
        "last_date": hist[-1]["date"],
        "dates": [h["date"] for h in hist],
    }


# ----------------------------------------------------------------------------
# 5) 输出：JSON / MD / HTML
# ----------------------------------------------------------------------------
def save_outputs(snapshot):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    d = snapshot["data_date"]

    latest = os.path.join(OUTPUT_DIR, "momentum_latest.json")
    with open(latest, "w", encoding="utf-8") as f:
        json.dump(snapshot, f, ensure_ascii=False, indent=2, default=str)
    hist = os.path.join(OUTPUT_DIR, "market_momentum_" + d + ".json")
    with open(hist, "w", encoding="utf-8") as f:
        json.dump(snapshot, f, ensure_ascii=False, indent=2, default=str)

    ts_path = append_timeseries(snapshot)

    md = _to_markdown(snapshot)
    md_path = os.path.join(OUTPUT_DIR, "market_momentum_" + d + ".md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md)

    html = _to_html(snapshot)
    html_path = os.path.join(OUTPUT_DIR, "market_momentum_" + d + ".html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)

    return {"latest": latest, "history": hist, "timeseries": ts_path, "md": md_path, "html": html_path}


def _to_markdown(s):
    rs = s["rotation_signal"]
    ts = timeseries_summary()
    L = []
    L.append("# 全市场动量观察（外部参考·非系统决策）")
    L.append("- 数据日期：" + s["data_date"] + " ｜ 宇宙规模：" + str(s["universe_size"]) + " 只 ｜ 基准：SPY")
    L.append("- 主模型 Model A = " + _formula_str("A") + "（相对动量百分位，只排名不预测）")
    L.append("- Model B（加 5D）= " + _formula_str("B") + "（RS5 由全宇宙 5 日超额收益百分位独立计算）")
    L.append("- 时间序列累积：" + (str(ts["snapshots"]) + " 次快照（" + ts["first_date"] + " ~ " + ts["last_date"] + "）" if ts else "首次"))
    L.append("- 档位分布：" + str(s["tier_distribution"]))
    L.append("\n## 资产轮动信号（Model A）")
    L.append("- 动量领先板块(二级)：" + "、".join(k + "(" + str(v) + ")" for k, v in rs["leading_sectors"]))
    L.append("- 动量落后板块(二级)：" + "、".join(k + "(" + str(v) + ")" for k, v in rs["lagging_sectors"]))
    L.append("- 动量领先资产类(一级)：" + "、".join(k + "(" + str(v) + ")" for k, v in rs["leading_asset_classes"]))
    L.append("- 动量落后资产类(一级)：" + "、".join(k + "(" + str(v) + ")" for k, v in rs["lagging_asset_classes"]))
    L.append("\n## 动量领先 Top15（Model A 综合排名）")
    for i, r in enumerate(s["top15"], 1):
        L.append(str(i) + ". " + r["ticker"] + " " + str(r["cn"]) + " ｜ " + r["cat2"] +
                 " ｜ A=" + str(r["composite"]) + " B=" + str(r["compB"]) +
                 " Tilt=" + str(r.get("tilt")) +
                 " ｜ RS(5/20/60/120)=" + str(r["rs5"]) + "/" + str(r["rs20"]) + "/" +
                 str(r["rs60"]) + "/" + str(r["rs120"]) + " ｜ 120日超额" + str(r["ex120"]))
    tv = s.get("tilt_view")
    if tv:
        L.append("\n## Momentum Tilt 观察（Tilt = Model B − Model A，研究阈值非信号）")
        L.append("- 阈值：> +10 短期显著增强 ｜ +5~+10 短期改善 ｜ −5~+5 基本稳定 ｜ −10~−5 短期走弱 ｜ < −10 短期明显恶化")
        L.append("- 短期动能最强（Tilt 高 → 趋势修复/加速）：")
        for x in tv["strongest_up"][:6]:
            L.append("    " + x["ticker"] + " " + str(x["cn"]) + "：A=" + str(x["compA"]) + " B=" + str(x["compB"]) +
                     " Tilt=" + str(x["tilt"]) + "（" + x["cls"] + "）")
        L.append("- 短期动能最弱（Tilt 低 → 趋势衰减）：")
        for x in tv["strongest_down"][:6]:
            L.append("    " + x["ticker"] + " " + str(x["cn"]) + "：A=" + str(x["compA"]) + " B=" + str(x["compB"]) +
                     " Tilt=" + str(x["tilt"]) + "（" + x["cls"] + "）")
    if s["model_sensitivity"]:
        L.append("\n## Model A vs B 排名敏感性（|Δrank|≥15）")
        for x in s["model_sensitivity"][:10]:
            L.append("- " + x["ticker"] + "：A 第" + str(x["rankA"]) + " → B 第" + str(x["rankB"]) +
                     "（" + ("B更弱" if x["delta"] > 0 else "B更强") + "）")
    L.append("\n## 建议观察池（相对强+绝对上行）")
    L.append("、".join(r["ticker"] for r in s["watch_pool"][:15]) or "（无）")
    L.append("\n## 回避池（弱势/绝对下行）")
    L.append("、".join(r["ticker"] for r in s["avoid_pool"][:15]) or "（无）")
    L.append("\n> 复刻校验：Model A composite 重算漂移 max=" + str(s["replication_drift"]["max"]) +
             " / min=" + str(s["replication_drift"]["min"]) + "（应~0）")
    L.append("> Model C(>200DMA)/D(动量÷波动)/E(Regime调制) 待价格序列/波动率/Regime 信号接入，当前未计算。")
    return "\n".join(L)


def _to_html(s):
    rs = s["rotation_signal"]
    ts = timeseries_summary()

    def row_top(r):
        return ("<tr><td>" + r["ticker"] + "</td><td>" + str(r["cn"]) + "</td><td>" + r["cat2"] +
                "</td><td class='num'>" + str(r["composite"]) + "</td><td class='num'>" +
                str(r["compB"]) + "</td><td class='num tilt" + _tilt_sign(r.get("tilt")) + "'>" +
                str(r.get("tilt")) + "</td><td class='num'>" + str(r["rs20"]) + "</td><td class='num'>" +
                str(r["rs60"]) + "</td><td class='num'>" + str(r["rs120"]) + "</td><td class='num'>" +
                str(r["ex120"]) + "</td></tr>")

    def _tilt_sign(v):
        if v is None:
            return ""
        return " pos" if v > 5 else (" neg" if v < -5 else "")

    top_rows = "".join(row_top(r) for r in s["top15"])
    lead = "".join("<span class='pill green'>" + k + " " + str(v) + "</span>" for k, v in rs["leading_sectors"])
    lag = "".join("<span class='pill red'>" + k + " " + str(v) + "</span>" for k, v in rs["lagging_sectors"])
    lead1 = "".join("<span class='pill green'>" + k + " " + str(v) + "</span>" for k, v in rs["leading_asset_classes"])
    lag1 = "".join("<span class='pill red'>" + k + " " + str(v) + "</span>" for k, v in rs["lagging_asset_classes"])
    watch = "、".join(r["ticker"] for r in s["watch_pool"][:15]) or "（无）"
    avoid = "、".join(r["ticker"] for r in s["avoid_pool"][:15]) or "（无）"
    drift = s["replication_drift"]
    ts_html = ("<div>时间序列累积：<b>" + str(ts["snapshots"]) + "</b> 次快照（" +
               ts["first_date"] + " ~ " + ts["last_date"] + "），构成未来回测底座</div>") if ts else ""
    sens = ""
    if s["model_sensitivity"]:
        items = "".join("<li>" + x["ticker"] + "：A第" + str(x["rankA"]) + "→B第" + str(x["rankB"]) +
                        "（" + ("B更弱" if x["delta"] > 0 else "B更强") + "）</li>" for x in s["model_sensitivity"][:8])
        sens = "<div class='card'><h2>Model A vs B 排名敏感性（|Δrank|≥15）</h2><ul class='sens'>" + items + "</ul></div>"

    # Tilt 观察卡片
    tv = s.get("tilt_view")
    tilt_card = ""
    if tv:
        up_items = "".join("<li>" + x["ticker"] + " " + str(x["cn"]) + "：A=" + str(x["compA"]) +
                           " B=" + str(x["compB"]) + " <b>Tilt=" + str(x["tilt"]) + "</b>（" + x["cls"] + "）</li>"
                           for x in tv["strongest_up"]) or "<li>（无，Tilt≤+5）</li>"
        dn_items = "".join("<li>" + x["ticker"] + " " + str(x["cn"]) + "：A=" + str(x["compA"]) +
                           " B=" + str(x["compB"]) + " <b>Tilt=" + str(x["tilt"]) + "</b>（" + x["cls"] + "）</li>"
                           for x in tv["strongest_down"]) or "<li>（无，Tilt≥−5）</li>"
        tilt_card = ("<div class='card'><h2>Momentum Tilt 观察（Tilt = Model B − Model A）</h2>"
                     "<div class='note'>研究观察阈值，非交易信号：&gt;+10 显著增强｜+5~+10 改善｜−5~+5 稳定｜−10~−5 走弱｜&lt;−10 明显恶化。预测价值待 P2（≥20 周度快照）验证。</div>"
                     "<h3 style='font-size:13px;margin:12px 0 6px;color:#a8e6c4'>短期动能最强（趋势修复 / 加速）</h3><ul class='sens'>" + up_items + "</ul>"
                     "<h3 style='font-size:13px;margin:12px 0 6px;color:#f0b9b9'>短期动能最弱（趋势衰减）</h3><ul class='sens'>" + dn_items + "</ul></div>")
    return """<!doctype html><html lang="zh"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>全市场动量观察 """ + s["data_date"] + """</title>
<style>
:root{--bg:#0d1410;--bg2:#16241c;--green:#2E7D52;--line:#26402f;--txt:#dfeee5;--mut:#8fae9c;}
*{box-sizing:border-box}
body{margin:0;background:linear-gradient(160deg,#0d1410,#16241c);color:var(--txt);
font-family:-apple-system,'Segoe UI',Roboto,'PingFang SC','Microsoft YaHei',sans-serif;padding:28px;}
.wrap{max-width:980px;margin:0 auto}
h1{font-size:22px;margin:0 0 4px}
.sub{color:var(--mut);font-size:13px;margin-bottom:18px}
.card{background:rgba(255,255,255,.03);border:1px solid var(--line);border-radius:12px;padding:16px 18px;margin-bottom:16px}
.card h2{font-size:15px;margin:0 0 10px;color:#7fd6a3}
.pill{display:inline-block;background:rgba(46,125,82,.18);border:1px solid var(--green);
border-radius:999px;padding:4px 11px;margin:3px;font-size:13px}
.pill.red{background:rgba(180,60,60,.16);border-color:#a84545;color:#f0b9b9}
.pill.green{color:#a8e6c4}
table{width:100%;border-collapse:collapse;font-size:13px}
th,td{text-align:left;padding:7px 8px;border-bottom:1px solid var(--line)}
th{color:var(--mut);font-weight:600}
td.num{text-align:right;font-variant-numeric:tabular-nums}
.note{color:var(--mut);font-size:12px;margin-top:8px}
.tag{color:#7fd6a3;font-weight:600}
.sens{margin:0;padding-left:18px;color:var(--mut);font-size:13px}
td.tilt.pos{color:#a8e6c4;font-weight:600}
td.tilt.neg{color:#f0b9b9;font-weight:600}
</style></head><body><div class="wrap">
<h1>全市场动量观察</h1>
<div class="sub">数据日期 """ + s["data_date"] + """ ｜ 宇宙 """ + str(s["universe_size"]) + """ 只 ｜ 基准 SPY ｜ 外部参考·非系统决策</div>

<div class="card"><h2>资产轮动信号（Model A）</h2>
<div>动量领先板块：""" + lead + """</div><div style="margin-top:6px">动量落后板块：""" + lag + """</div>
<div style="margin-top:10px">领先资产类：""" + lead1 + """</div><div style="margin-top:6px">落后资产类：""" + lag1 + """</div>
""" + ts_html + """
</div>

<div class="card"><h2>动量领先 Top15（Model A 综合排名）</h2>
<table><thead><tr><th>代码</th><th>名称</th><th>二级分类</th><th>综合A</th><th>综合B</th><th>Tilt(B-A)</th><th>RS20</th><th>RS60</th><th>RS120</th><th>120日超额</th></tr></thead>
<tbody>""" + top_rows + """</tbody></table>
<div class="note">Model A = """ + _formula_str("A") + """ ｜ Model B = """ + _formula_str("B") + """（RS5 由全宇宙 5 日超额收益百分位独立计算）</div>
</div>

""" + sens + tilt_card + """

<div class="card"><h2>观察池 / 回避池</h2>
<div><span class="tag">建议观察</span>（相对强+绝对上行）：""" + watch + """</div>
<div style="margin-top:8px"><span class="tag">回避</span>（弱势/绝对下行）：""" + avoid + """</div>
</div>

<div class="card"><h2>复刻校验</h2>
<div>Model A composite 重算漂移 max=""" + str(drift["max"]) + """ / min=""" + str(drift["min"]) + """（应~0，验证摄入与源表公式一致）</div>
<div class="note">Model C(>200DMA)/D(动量÷波动)/E(Regime调制) 待数据接入，当前未计算。本模块为外部观察信号，不进入生产决策链，不修改 risk_guard / shadow / run_daily。</div>
</div>
</div></body></html>"""


# ----------------------------------------------------------------------------
# main
# ----------------------------------------------------------------------------
def _resolve_path(arg_path):
    if arg_path and os.path.exists(arg_path):
        return arg_path
    candidates = [
        arg_path,
        os.path.join(IMPORTS_DIR, "market_momentum_latest.csv"),
        os.path.join(IMPORTS_DIR, "market_momentum_latest.xlsx"),
        DEFAULT_XLSX,
    ]
    for c in candidates:
        if c and os.path.exists(c):
            return c
    return arg_path


def main():
    arg = sys.argv[1] if len(sys.argv) > 1 else None
    path = _resolve_path(arg)
    if not path or not os.path.exists(path):
        print("[ERROR] 文件不存在: " + str(path))
        sys.exit(1)
    print("[1/3] 读取: " + path)
    parsed = parse_any(path)
    tabs = {k: len(v) for k, v in parsed.items()}
    print("      数据表: " + str(tabs))

    print("[2/3] 构建快照 + 双模型 + 复刻校验")
    snapshot = build_snapshot(parsed)

    print("[3/3] 写出（含时间序列累积）")
    paths = save_outputs(snapshot)
    for k, v in paths.items():
        print("      " + k + ": " + v)
    ts = timeseries_summary()
    print("\n完成。数据日期=" + snapshot["data_date"] + " 宇宙=" + str(snapshot["universe_size"]) +
          " 漂移 max=" + str(snapshot["replication_drift"]["max"]) + " min=" + str(snapshot["replication_drift"]["min"]) +
          " 时间序列快照数=" + str(ts["snapshots"] if ts else 1))


if __name__ == "__main__":
    main()
