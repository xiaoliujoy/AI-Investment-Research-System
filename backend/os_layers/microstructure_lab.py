"""
microstructure_lab.py - 加密市场微观结构实验室（研究用，非 CIO 模块）

来源：交易猫「市场数据终端.xlsx」（Binance 加密货币微观结构异动终端）。
定位：本脚本是一个「实验室数据集」生成器，把终端的多层字段结构化，
      并按 Price × Positioning × Flow 框架做实验性状态分类。
      ⚠️ 它不进入 AI CIO 生产决策链，不修改 risk_guard / shadow / run_daily。
         它属于用户的「Market Microstructure Lab」研究计划（P3），
         用途是未来验证「OI+价格+主动资金」组合是否真有增量 Alpha。

====================================================================
终端的四层结构（从表字段还原）
====================================================================
  ① Price         : 5m/15m/1h/4h/1d/1w 价变、波动、EMA、趋势、区间、breakout
  ② Positioning   : 持仓量/持仓额变化、Funding、大户持仓比、市场账户、结构分歧分
  ③ Flow          : 主动买卖比/差、CVD、spot/futures imbalance、资金流向
  ④ Price×Position: 价涨+OI涨(新增仓位推涨) vs 价涨+OI跌(空头回补) 等组合状态

====================================================================
实验性四态分类（仅研究假设，未经统计验证，禁止当信号用）
====================================================================
  Trend Expansion  : 价↑ & OI↑ & 主动买↑   —— 新增仓位推动上涨
  Short Covering   : 价↑ & OI↓ & 主动买↑   —— 空头回补推动上涨
  Short Expansion  : 价↓ & OI↑ & 主动买↓   —— 空头建仓压制
  Long Liquidation : 价↓ & OI↓ & 主动买↑   —— 多头清算
  Mixed/Other      : 其余

====================================================================
用法
====================================================================
  python microstructure_lab.py [市场数据终端xlsx路径]
  默认路径：D:/Downloads/市场数据终端.xlsx
输出：
  backend/output/microstructure_lab_latest.json
  backend/output/microstructure_lab_<日期>.md
"""

import os
import sys
import json
import datetime as dt
from collections import defaultdict, Counter

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.abspath(os.path.join(HERE, "..", "output"))
DEFAULT_XLSX = r"D:/Downloads/市场数据终端.xlsx"

ANOMALY_SHEET = "异动面板"
STAT_SHEET = "全市场统计"
DESC_SHEET = "市场描述"


def _find_idx(header, name):
    for i, h in enumerate(header):
        if h == name:
            return i
    return None


def _num(x):
    try:
        return float(x) if x is not None else None
    except (TypeError, ValueError):
        return None


def _load_sheet(path, sheet):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return None, None
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = None
    start = 0
    for i, r in enumerate(rows):
        if any(c is not None for c in r):
            hdr = list(r)
            start = i + 1
            break
    return hdr, rows[start:]


def parse_anomaly(path):
    hdr, rows = _load_sheet(path, ANOMALY_SHEET)
    if hdr is None:
        return []
    want = {
        "排名": "rank", "交易对": "pair", "综合分": "score", "方向": "direction",
        "强度等级": "strength", "主异动类型": "anomaly_type", "主时间窗口": "window",
        "摘要": "summary", "命中数": "hit_count", "结构形态": "structure",
        "清仓/挤压风险": "squeeze_risk", "主指标": "main_metric", "主指标值": "main_metric_val",
        "现持仓额": "oi_value", "最新价格": "price", "当前份额(%)": "share_pct",
        "主动买卖比": "active_buy_sell_ratio", "主动买卖差": "active_buy_sell_diff",
        "大户持仓比": "whale_ratio", "大户仓位-市场账户": "whale_gap",
        "结构分歧分": "structure_divergence", "波动强度": "vol_strength",
        "持仓量变化": "oi_change", "持仓额变化": "oi_value_change",
        "5m价变(%)": "chg_5m", "15m价变(%)": "chg_15m", "1h价变(%)": "chg_1h",
        "4h价变(%)": "chg_4h", "1d价变(%)": "chg_1d", "1w价变(%)": "chg_1w",
        "5m额强度(z)": "amt_z_5m", "15m额强度(z)": "amt_z_15m", "1h额强度(z)": "amt_z_1h",
        "4h额强度(z)": "amt_z_4h", "1d额强度(z)": "amt_z_1d", "1w额强度(z)": "amt_z_1w",
        "数据新鲜度分": "freshness", "更新时间": "update_time",
    }
    idx = {src: _find_idx(hdr, src) for src in want}
    out = []
    for r in rows:
        if len(r) <= 1 or r[0] in (None, ""):
            continue
        rec = {}
        for src, key in want.items():
            i = idx.get(src)
            if i is None or i >= len(r):
                rec[key] = None
                continue
            v = r[i]
            if key in ("rank", "pair", "direction", "strength", "anomaly_type", "window",
                       "summary", "structure", "squeeze_risk", "main_metric", "update_time"):
                rec[key] = str(v) if v is not None else None
            else:
                rec[key] = _num(v)
        rec["lab_state"] = classify_state(rec)
        out.append(rec)
    return out


def classify_state(rec):
    """实验性四态分类（研究假设，禁止当信号）。"""
    chg = rec.get("chg_1h") if rec.get("chg_1h") is not None else rec.get("chg_15m")
    oi_up = (rec.get("oi_change") or 0) > 0
    buy_up = (rec.get("active_buy_sell_ratio") or 0) > 1 or (rec.get("active_buy_sell_diff") or 0) > 0
    price_up = (chg or 0) > 0
    if price_up and oi_up and buy_up:
        return "Trend Expansion"
    if price_up and (not oi_up) and buy_up:
        return "Short Covering"
    if (not price_up) and oi_up and (not buy_up):
        return "Short Expansion"
    if (not price_up) and (not oi_up) and buy_up:
        return "Long Liquidation"
    return "Mixed/Other"


def parse_market_stat(path):
    hdr, rows = _load_sheet(path, STAT_SHEET)
    if hdr is None:
        return []
    keys = ["窗口", "覆盖合约数", "当前持仓额", "基线持仓额", "持仓额总变动", "持仓额总变动率",
            "持仓额净变化", "净变化占总变动(%)", "持仓额增加合约数", "持仓额减少合约数",
            "净增加广度(%)", "Top1持仓额增加交易对", "Top1持仓额减少交易对", "Top5增加占比", "Top5减少占比"]
    idx = {k: _find_idx(hdr, k) for k in keys}
    out = []
    for r in rows:
        if len(r) <= 1 or r[0] in (None, ""):
            continue
        rec = {}
        for k in keys:
            i = idx.get(k)
            v = r[i] if (i is not None and i < len(r)) else None
            rec[k] = str(v) if k in ("窗口", "Top1持仓额增加交易对", "Top1持仓额减少交易对") else _num(v)
        out.append(rec)
    return out


def parse_market_desc(path):
    hdr, rows = _load_sheet(path, DESC_SHEET)
    if hdr is None:
        return []
    keys = ["交易对", "产品", "数据时间(UTC+8)", "服务状态", "当前状态", "状态维度",
            "数据可信度", "覆盖率(%)", "完整率(%)", "新鲜度(s)", "质量状态", "关键证据", "异常", "冲突", "局限", "数据范围"]
    idx = {k: _find_idx(hdr, k) for k in keys}
    out = []
    for r in rows:
        if len(r) <= 1 or r[0] in (None, ""):
            continue
        rec = {}
        for k in keys:
            i = idx.get(k)
            v = r[i] if (i is not None and i < len(r)) else None
            rec[k] = str(v) if v is not None else None
        out.append(rec)
    return out


def build_lab(path):
    anomalies = parse_anomaly(path)
    stats = parse_market_stat(path)
    descs = parse_market_desc(path)

    # 四态分布（实验性）
    state_counter = Counter(r["lab_state"] for r in anomalies)
    # 异动强度分布
    by_score = sorted([r for r in anomalies if r.get("score") is not None],
                      key=lambda r: r["score"], reverse=True)

    lab = {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "source": "市场数据终端.xlsx（交易猫 Binance 加密微观结构终端）",
        "warning": "研究实验室数据集，非 CIO 信号；四态分类为实验性假设，未经统计验证",
        "coverage": {
            "anomaly_rows": len(anomalies),
            "market_stat_windows": len(stats),
            "market_desc_pairs": len(descs),
        },
        "lab_state_distribution": dict(state_counter),
        "aggregate_oi_flow": stats,           # 各窗口持仓额变动（市场级资金行为）
        "market_state_desc": descs,          # 每对 market_state + 资金行为状态
        "top_anomalies": [
            {k: r[k] for k in ("rank", "pair", "score", "direction", "strength",
                               "anomaly_type", "structure", "squeeze_risk", "lab_state",
                               "active_buy_sell_ratio", "whale_ratio", "chg_1h", "oi_change")}
            for r in by_score[:15]
        ],
        "records": anomalies,                # 全量，供实验室深挖
    }
    return lab


def save_outputs(lab):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    today = dt.date.today().isoformat()
    latest = os.path.join(OUTPUT_DIR, "microstructure_lab_latest.json")
    with open(latest, "w", encoding="utf-8") as f:
        json.dump(lab, f, ensure_ascii=False, indent=2, default=str)
    hist = os.path.join(OUTPUT_DIR, "microstructure_lab_" + today + ".json")
    with open(hist, "w", encoding="utf-8") as f:
        json.dump(lab, f, ensure_ascii=False, indent=2, default=str)
    md = _to_markdown(lab)
    md_path = os.path.join(OUTPUT_DIR, "microstructure_lab_" + today + ".md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md)
    return {"latest": latest, "history": hist, "md": md_path}


def _to_markdown(lab):
    L = []
    L.append("# 加密市场微观结构实验室（研究用·非 CIO 信号）")
    L.append("- 生成：" + lab["generated_at"] + " ｜ 覆盖：异动" + str(lab["coverage"]["anomaly_rows"]) +
             " 行 / 全市场统计" + str(lab["coverage"]["market_stat_windows"]) + " 窗口 / 描述" +
             str(lab["coverage"]["market_desc_pairs"]) + " 对")
    L.append("> ⚠️ " + lab["warning"])
    L.append("\n## 实验性四态分布（Price×OI×Flow 假设，未验证）")
    for k, v in lab["lab_state_distribution"].items():
        L.append("- " + k + "：" + str(v))
    L.append("\n## 全市场持仓额变动（各窗口，市场级资金行为）")
    for s in lab["aggregate_oi_flow"]:
        L.append("- " + str(s["窗口"]) + "：总变动率 " + str(s["持仓额总变动率"]) +
                 "% ｜ 净变化 " + str(s["持仓额净变化"]) + " ｜ 增加/减少合约 " +
                 str(s["持仓额增加合约数"]) + "/" + str(s["持仓额减少合约数"]) +
                 " ｜ Top1增=" + str(s["Top1持仓额增加交易对"]) + " Top1减=" + str(s["Top1持仓额减少交易对"]))
    L.append("\n## 异动 Top（综合分降序）")
    for r in lab["top_anomalies"]:
        L.append("- #" + str(r["rank"]) + " " + str(r["pair"]) + " 分=" + str(r["score"]) +
                 " 方向=" + str(r["direction"]) + " 结构=" + str(r["structure"]) +
                 " 态=" + str(r["lab_state"]) + " 主动买卖比=" + str(r["active_buy_sell_ratio"]) +
                 " 大户比=" + str(r["whale_ratio"]) + " 1h价变=" + str(r["chg_1h"]) + " OI变=" + str(r["oi_change"]))
    L.append("\n## 各交易对市场状态（market_state + 资金行为）")
    for d in lab["market_state_desc"][:20]:
        L.append("- " + str(d["交易对"]) + "：" + str(d["当前状态"]) + "（" + str(d["状态维度"]) +
                 "）可信度=" + str(d["数据可信度"]) + " 证据=" + str(d["关键证据"]))
    return "\n".join(L)


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(path):
        print("[ERROR] 文件不存在: " + path)
        sys.exit(1)
    print("[1/2] 读取: " + path)
    lab = build_lab(path)
    print("      异动" + str(lab["coverage"]["anomaly_rows"]) + " 行 / 统计" +
          str(lab["coverage"]["market_stat_windows"]) + " 窗口 / 描述" +
          str(lab["coverage"]["market_desc_pairs"]) + " 对")
    print("      四态分布: " + str(lab["lab_state_distribution"]))
    print("[2/2] 写出实验室数据集")
    paths = save_outputs(lab)
    for k, v in paths.items():
        print("      " + k + ": " + v)
    print("\n完成。实验室数据集已生成（非 CIO 信号）。")


if __name__ == "__main__":
    main()
