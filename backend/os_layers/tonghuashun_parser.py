# -*- coding: utf-8 -*-
"""
同花顺导出解析器（只读）
======================
- 股票账户-成交记录.xls  : 实为 GBK Tab 分隔文本(资金流水), 过滤 操作=买入/卖出 的真实股票成交
- 期货账户-成交记录.xlsx  : openpyxl 读取 '成交记录' 表(与文华账单同构)
输出归一化 CSV + 写入 vibe_research.db (tonghuashun_stock_trade / tonghuashun_futures_trade)
股票额外做 FIFO 配对, 输出 tonghuashun_stock_realized (逐股已实现盈亏)
"""
import argparse
import csv
import os
import sqlite3
import sys
from datetime import date

DB = os.path.join(os.path.dirname(__file__), "..", "..", "backend", "database", "vibe_research.db")
OUT = os.path.join(os.path.dirname(__file__), "..", "..", "mt5_raw")


def _f(x):
    if x is None:
        return None
    s = str(x).replace(",", "").strip()
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
def parse_stock_tsv(path):
    rows = []
    with open(path, encoding="gbk", errors="replace") as f:
        reader = csv.reader(f, delimiter="\t")
        header = next(reader)
        idx = {name: i for i, name in enumerate(header)}
        for raw in reader:
            if len(raw) <= idx.get("操作", -1):
                continue
            op = raw[idx["操作"]].strip()
            if op not in ("证券买入", "证券卖出"):
                continue
            code = raw[idx["证券代码"]].strip()
            if not code:
                continue
            name = raw[idx["证券名称"]].strip()
            tdate = raw[idx["成交日期"]].strip() or raw[idx["日期"]].strip()
            qty = _f(raw[idx["成交数量"]])
            price = _f(raw[idx["成交均价"]])
            amount = _f(raw[idx["成交金额"]])
            occ = _f(raw[idx["发生金额"]])
            fee = _f(raw[idx["手续费"]]) or 0
            tax = _f(raw[idx["印花税"]]) or 0
            oth = _f(raw[idx["其他杂费"]]) or 0
            acct = raw[idx["资金帐户"]].strip()
            market = raw[idx["交易市场"]].strip()
            rows.append({
                "trade_date": tdate, "code": code, "name": name, "op": op,
                "qty": qty, "price": price, "amount": amount, "occ": occ,
                "fee": fee, "stamp": tax, "other": oth, "account": acct,
                "market": market,
            })
    return rows


def fifo_realized(stock_rows):
    """按 证券代码 FIFO 配对, 返回 逐股已实现盈亏 list"""
    from collections import defaultdict, deque
    queues = defaultdict(deque)  # code -> deque of (price, qty, fee)
    results = defaultdict(lambda: {"buy_amt": 0.0, "sell_amt": 0.0,
                                    "fee": 0.0, "realized": 0.0, "buy_qty": 0, "sell_qty": 0,
                                    "name": "", "market": ""})
    for r in stock_rows:
        code = r["code"]
        res = results[code]
        res["name"] = r["name"]
        res["market"] = r["market"]
        qty = r["qty"] or 0
        price = r["price"] or 0
        fee = (r["fee"] or 0) + (r["stamp"] or 0) + (r["other"] or 0)
        if r["op"] == "证券买入":
            res["buy_amt"] += (price * qty)
            res["buy_qty"] += qty
            res["fee"] += fee
            queues[code].append([price, qty, fee])
        else:  # 证券卖出
            res["sell_qty"] += qty
            res["sell_amt"] += (price * qty)
            res["fee"] += fee
            remain = qty
            while remain > 0 and queues[code]:
                lot = queues[code][0]
                take = min(remain, lot[1])
                buy_fee_share = lot[2] * (take / lot[1]) if lot[1] else 0
                res["realized"] += (price - lot[0]) * take - buy_fee_share
                remain -= take
                lot[1] -= take
                if lot[1] <= 0:
                    queues[code].popleft()
            if remain > 0:  # 卖超买(分红/配股等), 差额计为盈亏
                res["realized"] += price * remain
    out = []
    for code, res in results.items():
        net = res["realized"] - res["fee"]
        out.append({"code": code, "name": res["name"], "market": res["market"],
                    "buy_qty": res["buy_qty"], "sell_qty": res["sell_qty"],
                    "buy_amt": round(res["buy_amt"], 2), "sell_amt": round(res["sell_amt"], 2),
                    "fee": round(res["fee"], 2), "realized": round(res["realized"], 2),
                    "net_pnl": round(net, 2)})
    return out


# ---------------------------------------------------------------------------
def parse_futures_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    trades = []
    cur_date = None
    for row in rows[1:]:
        if row[0] is not None and str(row[0]).strip() != "":
            cur_date = str(row[0]).strip()
            if row[2] is None or str(row[2]).strip() == "":
                continue  # 日期分隔行
        instr = row[2]
        if instr is None or str(instr).strip() == "":
            continue
        trades.append({
            "trade_date": cur_date,
            "exchange": str(row[1]).strip() if row[1] else "",
            "instrument": str(row[2]).strip(),
            "bs": str(row[3]).strip() if row[3] else "",
            "oc": str(row[4]).strip() if row[4] else "",
            "lots": _f(row[5]), "price": _f(row[6]), "turnover": _f(row[7]),
            "realized": _f(row[8]), "fee": _f(row[9]),
            "sh": str(row[10]).strip() if row[10] else "",
            "premium": _f(row[11]),
            "net_pnl": (_f(row[8]) or 0) + (_f(row[11]) or 0) - (_f(row[9]) or 0),
        })
    return trades


# ---------------------------------------------------------------------------
def save(db_table, cols, rows):
    con = sqlite3.connect(DB)
    con.execute("DROP TABLE IF EXISTS %s" % db_table)
    con.execute("CREATE TABLE %s (%s)" % (db_table, ", ".join("%s TEXT" % c if c in ("trade_date", "code", "name", "op", "account", "market", "exchange", "instrument", "bs", "oc", "sh") else "%s REAL" % c for c in cols)))
    placeholders = ",".join("?" * len(cols))
    con.executemany("INSERT INTO %s (%s) VALUES (%s)" % (db_table, ",".join(cols), placeholders),
                    [[r.get(c) for c in cols] for r in rows])
    con.commit()
    n = con.execute("SELECT COUNT(*) FROM %s" % db_table).fetchone()[0]
    con.close()
    return n


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--stock", default=r"C:\Users\JOY\Desktop\股票账户-成交记录.xls")
    ap.add_argument("--futures", default=r"C:\Users\JOY\Desktop\期货账户-成交记录.xlsx")
    args = ap.parse_args()

    # ---- 股票 ----
    if os.path.exists(args.stock):
        srows = parse_stock_tsv(args.stock)
        scols = ["trade_date", "code", "name", "op", "qty", "price", "amount",
                 "occ", "fee", "stamp", "other", "account", "market"]
        n = save("tonghuashun_stock_trade", scols, srows)
        realized = fifo_realized(srows)
        rcols = ["code", "name", "market", "buy_qty", "sell_qty", "buy_amt",
                 "sell_amt", "fee", "realized", "net_pnl"]
        save("tonghuashun_stock_realized", rcols, realized)
        os.makedirs(OUT, exist_ok=True)
        with open(os.path.join(OUT, "tonghuashun_stock_trades.csv"), "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=scols); w.writeheader(); w.writerows(srows)
        tot_net = sum(r["net_pnl"] for r in realized)
        print("[股票] 成交fills=%d | 涉及个股=%d | FIFO已实现净盈亏=%.2f" % (n, len(realized), tot_net))
        print("       前5大盈亏个股:")
        for r in sorted(realized, key=lambda x: x["net_pnl"])[:5] + sorted(realized, key=lambda x: -x["net_pnl"])[:5]:
            print("         %s %s %s 净盈亏=%+.2f" % (r["code"], r["name"], r["market"], r["net_pnl"]))
    else:
        print("[股票] 文件不存在: %s" % args.stock)

    # ---- 期货 ----
    if os.path.exists(args.futures):
        frows = parse_futures_xlsx(args.futures)
        fcols = ["trade_date", "exchange", "instrument", "bs", "oc", "lots",
                 "price", "turnover", "realized", "fee", "sh", "premium", "net_pnl"]
        n = save("tonghuashun_futures_trade", fcols, frows)
        os.makedirs(OUT, exist_ok=True)
        with open(os.path.join(OUT, "tonghuashun_futures_trades.csv"), "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=fcols); w.writeheader(); w.writerows(frows)
        dates = [r["trade_date"] for r in frows if r["trade_date"]]
        net = sum(r["net_pnl"] or 0 for r in frows)
        print("\n[期货-同花顺] 成交fills=%d | 日期 %s~%s | 净盈亏(初算)=%.2f" %
              (n, min(dates), max(dates), net))
        from collections import Counter
        prod = Counter()
        for r in frows:
            prod[r["instrument"][:2]] += 1
        print("       合约前缀分布:", dict(prod))
    else:
        print("[期货] 文件不存在: %s" % args.futures)


if __name__ == "__main__":
    main()
